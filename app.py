from flask import Flask, request, send_file
import pandas as pd
import re
import io
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

@app.route('/procesar', methods=['POST'])
def procesar():
    data = request.get_json()
    print('KEYS RECIBIDAS:', list(data.keys()) if data else 'SIN DATA')
    print('CONTENT TYPE:', request.content_type)
    
    mayor_bytes = base64.b64decode(data['Mayor_fumi'])

    # ── 1. MAYOR FUMI ─────────────────────────────────────────────────────────
    df_mayor = pd.read_excel(io.BytesIO(mayor_bytes), header=4)
    df_mayor['NUM_COMPROBANTE'] = pd.to_numeric(df_mayor['NUM_COMPROBANTE'], errors='coerce')

    def transform_num(row):
        comp = str(row['Comprobante']).strip()
        num = row['NUM_COMPROBANTE']
        if pd.isna(num): return num
        num_int = int(num)
        return int(f'100{num_int}') if comp.startswith('LD') else num_int

    df_mayor['NUM_KEY'] = df_mayor.apply(transform_num, axis=1)

    # ── 2. DETALLE ITEMS MANUALES ─────────────────────────────────────────────
    df_items = pd.read_excel(io.BytesIO(items_bytes), header=1)
    real_cols = df_items.iloc[0].tolist()
    df_items.columns = real_cols
    df_items = df_items.iloc[1:].reset_index(drop=True)
    df_items['Nro Comprobante'] = pd.to_numeric(df_items['Nro Comprobante'], errors='coerce')
    df_items['Suma de Neto']    = pd.to_numeric(df_items['Suma de Neto'], errors='coerce')

    def extract_ctg(concepto):
        if pd.isna(concepto): return None
        s = str(concepto)
        s_clean = re.sub(r'(\d{4})-(\d+)', r'\1\2', s)
        m = re.search(r'(?:CTG|ctg)[^\d]*([\d]{10,12})', s_clean)
        if m: return str(int(m.group(1)))
        m = re.search(r'(\d{10,12})(?:\s|$)', s_clean)
        if m: return str(int(m.group(1)))
        return None

    df_items['CTG_extracted'] = df_items['Concepto'].apply(extract_ctg)

    # ── 3. DETALLE APLICACIONES ───────────────────────────────────────────────
    df_aplic = pd.read_excel(io.BytesIO(aplic_bytes))
    df_aplic['CTG_str'] = df_aplic['CTG'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
    aplic_dict = {}
    for _, row in df_aplic.iterrows():
        ctg = row['CTG_str']
        if ctg and ctg not in aplic_dict:
            aplic_dict[ctg] = {'Vendedor': row['Vendedor'], 'Contrato': row['Contrato']}

    # ── 4. CONSTRUIR RESULTADO ────────────────────────────────────────────────
    rows = []
    for _, mayor_row in df_mayor.iterrows():
        comprobante = mayor_row['Comprobante']
        num_key     = mayor_row['NUM_KEY']
        items_match = df_items[df_items['Nro Comprobante'] == num_key]

        if items_match.empty:
            rows.append({'Comprobante': comprobante, 'NUM_COMPROBANTE': num_key,
                         'Concepto (CTG)': 'Sin datos en detalle items', 'CTG': None,
                         'Comprador': None, 'Importe Neto': None, 'Vendedor': None, 'Contrato': None})
        else:
            for _, item_row in items_match.iterrows():
                ctg = item_row['CTG_extracted']
                aplic_info = aplic_dict.get(ctg, {}) if ctg else {}
                rows.append({'Comprobante': comprobante, 'NUM_COMPROBANTE': num_key,
                             'Concepto (CTG)': item_row['Concepto'], 'CTG': ctg,
                             'Comprador': item_row['Comprador'], 'Importe Neto': item_row['Suma de Neto'],
                             'Vendedor': aplic_info.get('Vendedor'), 'Contrato': aplic_info.get('Contrato')})

    df_result = pd.DataFrame(rows)

    # ── 5. CREAR EXCEL ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Detalle Fumigaciones'

    HEADER_BG = 'FF1F4E79'
    HEADER_FG = 'FFFFFFFF'
    COMP_BG   = 'FFD6E4F0'
    ALT_BG    = 'FFEFF7FB'

    thin        = Side(style='thin', color='FFBFBFBF')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ['Comprobante','NUM_COMPROBANTE','Concepto (CTG)','CTG','Comprador','Importe Neto','Vendedor','Contrato']
    col_widths = [24, 18, 52, 16, 32, 15, 32, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(name='Arial', bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill('solid', start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    current_comp = None
    comp_toggle  = True

    for row_idx, (_, row) in enumerate(df_result.iterrows(), 2):
        if row['Comprobante'] != current_comp:
            current_comp = row['Comprobante']
            comp_toggle  = not comp_toggle
        bg = COMP_BG if comp_toggle else ALT_BG
        values = [row['Comprobante'], row['NUM_COMPROBANTE'], row['Concepto (CTG)'],
                  row['CTG'], row['Comprador'], row['Importe Neto'], row['Vendedor'], row['Contrato']]
        for col_idx, value in enumerate(values, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name='Arial', size=9)
            cell.fill      = PatternFill('solid', start_color=bg)
            cell.border    = border_thin
            cell.alignment = Alignment(vertical='center')
            if col_idx == 6 and value is not None:
                cell.number_format = '#,##0.00'
                cell.alignment     = Alignment(horizontal='right', vertical='center')
            if col_idx == 8 and value is not None:
                try: cell.value = int(value)
                except: pass

    ws.freeze_panes    = 'A2'
    ws.auto_filter.ref = f'A1:H{len(df_result)+1}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_b64 = base64.b64encode(output.read()).decode('utf-8')

    return {'file': excel_b64, 'filename': 'Detalle_Fumigaciones.xlsx'}

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
